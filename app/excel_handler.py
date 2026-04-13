"""
Utilitários para leitura e escrita de Excel.
"""

import io
import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from carrier_lookup import normalize_number, is_valid_number

logger = logging.getLogger(__name__)


def read_numbers_from_excel(content: bytes) -> tuple[list[str], list[str]]:
    """
    Lê números da primeira coluna do Excel (ignora cabeçalho se não for número).
    Retorna (numeros_validos, numeros_invalidos).
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    valid = []
    invalid = []
    seen = set()

    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        raw = row[0]
        if raw is None:
            continue

        normalized = normalize_number(raw)

        # Pula linha de cabeçalho (texto puro sem dígitos suficientes)
        if not normalized or len(normalized) < 8:
            continue

        if normalized in seen:
            continue
        seen.add(normalized)

        if is_valid_number(normalized):
            valid.append(normalized)
        else:
            invalid.append(normalized)

    wb.close()
    logger.info("Excel lido: %d válidos | %d inválidos", len(valid), len(invalid))
    return valid, invalid


def write_results_to_excel(
    numbers_ordered: list[str],
    results: dict[str, dict],
    invalid_numbers: list[str] | None = None,
) -> bytes:
    """
    Gera Excel de resultado com formatação profissional.
    Retorna os bytes do arquivo .xlsx.
    """
    wb = openpyxl.Workbook()

    # ── Aba principal ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Operadoras"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    fills = {
        "sucesso": PatternFill("solid", fgColor="E2EFDA"),
        "nao_encontrado": PatternFill("solid", fgColor="FFF2CC"),
        "erro": PatternFill("solid", fgColor="FCE4D6"),
        "default": PatternFill("solid", fgColor="FFFFFF"),
    }

    # Cabeçalho
    headers = ["#", "Número", "Operadora", "Status", "Método"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin

    ws.row_dimensions[1].height = 22

    # Dados
    for row_idx, number in enumerate(numbers_ordered, 2):
        result = results.get(
            number,
            {"numero": number, "operadora": "NÃO_PROCESSADO", "status": "ausente"},
        )

        status = result.get("status", "")
        fill_key = (
            "sucesso" if status == "sucesso"
            else "nao_encontrado" if status == "nao_encontrado"
            else "erro" if "erro" in status.lower() or "falha" in status.lower()
            else "default"
        )
        row_fill = fills[fill_key]

        values = [
            row_idx - 1,
            result.get("numero", number),
            result.get("operadora", ""),
            status,
            result.get("metodo", ""),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = thin
            cell.alignment = center if col == 1 else left

    # Largura das colunas
    col_widths = [6, 18, 25, 22, 18]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header
    ws.freeze_panes = "A2"

    # ── Aba de inválidos ───────────────────────────────────────
    if invalid_numbers:
        ws_inv = wb.create_sheet("Inválidos")
        ws_inv.cell(1, 1, "Número").font = Font(bold=True)
        ws_inv.cell(1, 2, "Motivo").font = Font(bold=True)
        for i, n in enumerate(invalid_numbers, 2):
            ws_inv.cell(i, 1, n)
            ws_inv.cell(i, 2, "Formato inválido (esperado 10-11 dígitos)")
        ws_inv.column_dimensions["A"].width = 18
        ws_inv.column_dimensions["B"].width = 40

    # ── Aba de resumo ──────────────────────────────────────────
    ws_sum = wb.create_sheet("Resumo")
    all_results = list(results.values())
    total = len(numbers_ordered)
    success = sum(1 for r in all_results if r.get("status") == "sucesso")
    not_found = sum(1 for r in all_results if r.get("status") == "nao_encontrado")
    errors = total - success - not_found

    summary_rows = [
        ("Total de números", total),
        ("Consultados com sucesso", success),
        ("Não encontrados", not_found),
        ("Erros/Falhas", errors),
        ("Inválidos (não consultados)", len(invalid_numbers) if invalid_numbers else 0),
    ]

    ws_sum.cell(1, 1, "Métrica").font = Font(bold=True)
    ws_sum.cell(1, 2, "Quantidade").font = Font(bold=True)
    for i, (label, value) in enumerate(summary_rows, 2):
        ws_sum.cell(i, 1, label)
        ws_sum.cell(i, 2, value)
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 15

    # ── Salvar em memória e retornar bytes ─────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
